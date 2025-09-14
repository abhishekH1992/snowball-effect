import os
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def generate_system_comments(invoice_details: Dict[str, List[Dict]], bucket_names: List[str]) -> str:
    """
    Generate system comments based on invoice details for each aging bucket.
    """
    comments = []
    
    for bucket_name in bucket_names:
        bucket_items = invoice_details.get(bucket_name, [])
        if bucket_items:
            comments.append(f"{bucket_name}:")
            for item in bucket_items:
                item_number = item.get('item_number', 'Unknown')
                item_id = item.get('item_id', 'Unknown')
                amount = item.get('amount', 0)
                is_negative = item.get('is_negative', False)
                item_type = item.get('item_type', 'invoice')
                
                # Format based on item type with ID
                if item_type == "credit_note":
                    comments.append(f"{item_number} (Credit Note, ID: {item_id}) = {amount:,.2f}")
                elif item_type == "overpayment":
                    comments.append(f"{item_number} (Overpayment, ID: {item_id}) = {amount:,.2f}")
                elif item_type == "bank_transaction":
                    comments.append(f"{item_number} (Bank Transaction, ID: {item_id}) = {amount:,.2f}")
                else:
                    if is_negative:
                        if item_number == "Invoice Overpayments":
                            comments.append(f"Invoice Overpayments (Paid upfront for future invoices) = {amount:,.2f}")
                        else:
                            comments.append(f"{item_number} (Credit/Overpayment, ID: {item_id}) = {amount:,.2f}")
                    else:
                        comments.append(f"{item_number} (Invoice, ID: {item_id}) = {amount:,.2f}")
            comments.append("")  # blank line between buckets
    
    return "\n".join(comments) if comments else ""


def export_report_to_excel(
    data: List[Dict[str, Any]],
    columns: List[Dict[str, str]],
    filename: str,
    sheet_name: str = "Report",
    title: Optional[str] = None,
    organization_name: Optional[str] = None,
    report_date: Optional[str] = None,
    output_dir: str = "tmp",
    include_totals: bool = True,
    include_percentages: bool = True
) -> str:
    """
    Export report data to Excel with styled headers, optional title/organization/date,
    and group separators only when the company changes.
    """
    os.makedirs(output_dir, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="366092")
    header_alignment = Alignment("center", "center")
    title_font = Font(bold=True, size=14)
    title_alignment = Alignment("center")
    border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin")
    )

    current_row = 1

    # Title, Org, Date
    if title:
        ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
        ws["A1"] = title
        ws["A1"].font = title_font
        ws["A1"].alignment = title_alignment
        current_row = 2

    if organization_name:
        ws.merge_cells(f"A{current_row}:{get_column_letter(len(columns))}{current_row}")
        ws[f"A{current_row}"] = organization_name
        ws[f"A{current_row}"].font = Font(bold=True, size=12)
        ws[f"A{current_row}"].alignment = Alignment("center")
        current_row += 1

    if report_date:
        ws.merge_cells(f"A{current_row}:{get_column_letter(len(columns))}{current_row}")
        ws[f"A{current_row}"] = report_date
        ws[f"A{current_row}"].font = Font(size=11)
        ws[f"A{current_row}"].alignment = Alignment("center")
        current_row += 1

    # Subtitle
    ws.merge_cells(f"A{current_row}:{get_column_letter(len(columns))}{current_row}")
    ws[f"A{current_row}"] = "Ageing by due date"
    ws[f"A{current_row}"].font = Font(size=11)
    ws[f"A{current_row}"].alignment = Alignment("center")
    current_row += 1

    # Headers
    for idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=current_row, column=idx, value=col["header"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        ws.column_dimensions[get_column_letter(idx)].width = col.get("width", 15)
    header_row = current_row  # Save the header row index

    # Sort data
    data = sorted(
        data,
        key=lambda r: (str(r.get("Company", "")).lower(), str(r.get("Contact", "")).lower())
    )

    # last_company = None

    for idx, row_data in enumerate(data):
        # Write the data row
        current_row += 1

        # Write row values
        for col_idx, col in enumerate(columns, start=1):
            val = row_data.get(col["key"], "")
            cell = ws.cell(row=current_row, column=col_idx)

            fmt = col.get("format")
            if fmt == "currency" and isinstance(val, (int, float)):
                cell.number_format = '"$"#,##0.00'
                cell.value = val
            elif fmt == "percentage" and isinstance(val, (int, float)):
                cell.number_format = '0.00%'
                cell.value = val/100 if val > 1 else val
            elif fmt == "date" and hasattr(val, "strftime"):
                cell.value = val
                cell.number_format = 'dd/mm/yyyy'
            elif fmt == "number" and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
                cell.value = val
            else:
                cell.value = val

            # System Comments styling
            if col["key"] == "System Comments" and isinstance(val, str) and val:
                cell.alignment = Alignment("left", "top", wrap_text=True)
                cell.font = Font(size=10)
                ws.row_dimensions[current_row].height = max(60, len(val.split("\n"))*15)
            else:
                cell.alignment = Alignment("left", "center")

            cell.border = border

    # Totals row
    if include_totals and data:
        totals_row = current_row + 1
        current_row = totals_row
        # calculate sums
        sums = {}
        for col in columns:
            key = col["key"]
            if key not in ("Contact", "Comments", "System Comments"):
                sums[key] = sum(r.get(key, 0) for r in data if isinstance(r.get(key), (int, float)))

        for idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=totals_row, column=idx)
            if col["key"] == "Business Unit":
                cell.value = "Total"
                cell.font = Font(bold=True)
            elif col["key"] in ("Company", "Contact", "Comments", "System Comments"):
                cell.value = ""
            else:
                v = sums.get(col["key"], 0)
                cell.value = v
                cell.font = Font(bold=True)
                if col.get("format") == "currency":
                    cell.number_format = '"$"#,##0.00'
                elif col.get("format") == "percentage":
                    cell.number_format = '0.00%'
                elif col.get("format") == "number":
                    cell.number_format = '#,##0.00'
            cell.border = border
            cell.alignment = Alignment("left", "center")

    # Percentages row
    if include_percentages and data and include_totals:
        pct_row = current_row + 1
        # reuse sums["Total"] if present
        grand = sums.get("Total", 0)
        for idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=pct_row, column=idx)
            if col["key"] == "Business Unit":
                cell.value = "Percentage"
                cell.font = Font(bold=True)
            elif col["key"] in ("Comments", "System Comments"):
                cell.value = ""
            elif col["key"] == "Total":
                cell.value = 1.0
                cell.number_format = '0.00%'
                cell.font = Font(bold=True)
            else:
                pct = (sums.get(col["key"], 0) / grand) if grand else 0
                cell.value = pct
                cell.number_format = '0.00%'
                cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = Alignment("left", "center")

    # Auto‐adjust row heights
    for row in ws.iter_rows():
        h = max(
            15,
            max(
                (len(str(c.value)) * 0.8) if c.value else 0
                for c in row
            )
        )
        ws.row_dimensions[row[0].row].height = min(50, h)

    # Turn on filter arrows across your entire header+data block
    first_header_row = header_row
    last_data_row = current_row
    last_data_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A{first_header_row}:{last_data_col}{last_data_row}"
    ws.freeze_panes = ws[f"A{first_header_row+1}"]

    # Save
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(output_dir, f"{filename}_{timestamp}.xlsx")
    wb.save(path)
    return path


def format_currency(value: float) -> str:
    return f"${value:,.2f}"


def format_percentage(value: float) -> str:
    return f"{value:.2f}%"
